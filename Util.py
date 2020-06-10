import tabula
from os import listdir
from os.path import isfile, join
import pandas as pd
import openpyxl
from time import clock
import random
import math
import Bio as bio
from Bio import SeqIO
import glob

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    def read_tb_txt(self, path):
        result_list = []
        with open(path, "r") as f:
            f.readline().replace("\n", "")
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == "":
                    break

                tmp_arr = tmp_line.split("\t")
                result_list.append(tmp_arr)

        return result_list

    def make_excel(self, path, result_list):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='final_idx')
        sheet.cell(row=row, column=3, value='NGS read')
        sheet.cell(row=row, column=4, value='In_Del_Sub')
        sheet.cell(row=row, column=5, value='Ref seq')

        for val_arr in result_list:
            row += 1
            sheet.cell(row=row, column=1, value=str(row - 1))
            sheet.cell(row=row, column=2, value=val_arr[0])
            sheet.cell(row=row, column=3, value=val_arr[1])
            sheet.cell(row=row, column=4, value=val_arr[2])
            sheet.cell(row=row, column=5, value=val_arr[3])
            sheet.cell(row=row, column=6, value=len(val_arr[1]))
            sheet.cell(row=row, column=7, value=len(val_arr[2]))
            sheet.cell(row=row, column=8, value=len(val_arr[3]))
            if len(val_arr[1]) == len(val_arr[2]) == len(val_arr[3]):
                sheet.cell(row=row, column=9, value="O")

        workbook.save(filename=path + self.ext_xlsx)


